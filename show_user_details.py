import tkinter as tk
from tkinter import ttk, filedialog, messagebox, font
import openpyxl
import os
import webbrowser
import base64
from datetime import datetime
from PIL import Image, ImageTk
import zipfile
import subprocess

class ShowUserDetails:
    def __init__(self, master, db):
        self.master = master
        self.db = db
        self.create_ui()

    def create_ui(self):
        
        main_container = tk.Canvas(self.master, bg="#ffffff", highlightthickness=0)
        main_container.pack(fill='both', expand=True)

        
        # try:
        #     bg_image_path = "assets/bg.jpg"  # Replace with the path to your background image
        #     bg_image = Image.open(bg_image_path)
        #     bg_image = bg_image.resize(
        #         (self.master.winfo_screenwidth(), self.master.winfo_screenheight()), 
        #         Image.LANCZOS  # Use LANCZOS for high-quality resizing
        #     )
        #     self.bg_image_tk = ImageTk.PhotoImage(bg_image)  # Keep a reference to avoid garbage collection
        #     main_container.create_image(0, 0, image=self.bg_image_tk, anchor="nw")
        # except Exception as e:
        #     print(f"Error loading background image: {e}")
        
        def setup_background():
            main_container.update()
            try:
                bg_image = Image.open("assets/bg.jpg")
                width = self.master.winfo_width()
                height = self.master.winfo_height()
                
                bg_image = bg_image.resize((width, height), Image.LANCZOS)
                self.bg_image_tk = ImageTk.PhotoImage(bg_image)
                main_container.create_image(0, 0, image=self.bg_image_tk, anchor="nw")
            except Exception as e:
                print(f"Error: {e}")

        def update_background(event):
            setup_background()

        setup_background()
        main_container.bind('<Configure>', update_background)

        
        title_label = tk.Label(main_container, text="Student Details", 
                               font=("Arial", 20, "bold"), bg="#edf5f8")
        title_label.pack(pady=10)

        
        tree_frame = tk.Frame(main_container, bg="#BDDDFC")
        tree_frame.pack(fill='both', expand=True, padx=20, pady=10)

        
        vertical_scrollbar = ttk.Scrollbar(tree_frame)
        vertical_scrollbar.pack(side='right', fill='y')

        horizontal_scrollbar = ttk.Scrollbar(tree_frame, orient='horizontal')
        horizontal_scrollbar.pack(side='bottom', fill='x')

        self.user_tree = ttk.Treeview(tree_frame, 
                                  columns=(
                                      "Serial", "First Name", "Middle Name", "Last Name", 
                                      "Age", "Gender", "Contact", "Email", "Address",
                                      "SSC %", "Education Type", "HSC/Diploma %", 
                                      "Diploma Stream", "Degree Status", "Degree Type", "Degree Stream", 
                                      "Degree %", "Current Year", "Completion Year", "Skills", "Resume",
                                      "Role", "Duration", "Joining Date", 
                                      "Ending Date", "Internship Type", "Stipend Amount", 
                                      "Stipend Frequency"
                                  ), 
                                  show='headings',
                                  yscrollcommand=vertical_scrollbar.set,
                                  xscrollcommand=horizontal_scrollbar.set)

        vertical_scrollbar.config(command=self.user_tree.yview)
        horizontal_scrollbar.config(command=self.user_tree.xview)
        
        self.user_tree.bind_all("<MouseWheel>", lambda e: self.user_tree.yview_scroll(-1 * int(e.delta/120), "units"))

        column_config = [
            ("Serial", 50, 50),              # (name, width, minwidth)
            ("First Name", 100, 80),
            ("Middle Name", 100, 80),
            ("Last Name", 100, 80),
            ("Age", 50, 50),
            ("Gender", 80, 60),
            ("Contact", 100, 80),
            ("Email", 150, 100),
            ("Address", 200, 150),
            ("SSC %", 80, 60),
            ("Education Type", 100, 80),
            ("HSC/Diploma %", 100, 80),
            ("Diploma Stream", 120, 100),
            ("Degree Status", 100, 80),
            ("Degree Type", 100, 80),
            ("Degree Stream", 120, 100),
            ("Degree %", 80, 60),
            ("Current Year", 100, 80),
            ("Completion Year", 100, 80),
            ("Skills", 200, 150),
            ("Resume", 80, 80),
            ("Role", 100, 80),
            ("Duration", 100, 80),
            ("Joining Date", 100, 80),
            ("Ending Date", 100, 80),
            ("Internship Type", 100, 80),
            ("Stipend Amount", 100, 80),
            ("Stipend Frequency", 100, 80),
        ]

        for idx, (col, width, minwidth) in enumerate(column_config):
            self.user_tree.heading(col, text=col, anchor='center')
            self.user_tree.column(col, 
                                width=width, 
                                minwidth=minwidth,  
                                stretch=False,       
                                anchor='center')     

        self.user_tree.tag_configure("view_pdf", foreground="blue", font=('Arial', 10, 'underline'))

        self.user_tree.pack(fill='both', expand=True)

        self.user_tree.bind('<ButtonRelease-1>', self.handle_click)

        tree_frame.pack_configure(fill='both', expand=True)

        self.load_combined_data()

        # export_btn = tk.Button(main_container, text="Export All Data to Excel", 
        #                     command=self.export_all_to_excel,
        #                     bg="#05d7ff", fg="BLACK", 
        #                     highlightthickness=2, 
        #                     cursor="hand1", 
        #                     highlightcolor="#05d7ff",
        #                     font=("Arial", 12))
        # export_btn.pack(pady=10)
        export_icon = Image.open("assets/export-btn-img.png").resize((100, 50))
        export_img = ImageTk.PhotoImage(export_icon)
        export_btn = tk.Button(main_container, image=export_img, bd=0, highlightthickness=0, borderwidth=0, command=self.export_all_to_excel)
        export_btn.image = export_img
        export_btn.pack(pady=8)

    def handle_click(self, event):
        region = self.user_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
            
        column = self.user_tree.identify_column(event.x)
        item = self.user_tree.identify_row(event.y)
        
        if not item:
            return
            
        col_num = int(column.replace('#', '')) - 1
        
        if col_num == 20:  
            cell_value = self.user_tree.item(item)['values'][col_num]
            
            if cell_value == "View PDF":
                try:
                    serial_no = self.user_tree.item(item)['values'][0]  
                    user = list(self.db.users_collection.find())[serial_no - 1]  
                    user_id = user['_id']  
                    
                    education_record = self.db.education_collection.find_one({'user_id': user_id})
                    if education_record and education_record.get('resume'):
                        self.open_resume_from_education(education_record)
                    else:
                        messagebox.showinfo("Resume", "No resume available for this user.")
                except Exception as e:
                    messagebox.showerror("Error", f"Could not retrieve resume: {str(e)}")

    def format_date(self, date_str):
        """Convert date string to a clean date format."""
        if not date_str or date_str == 'N/A':
            return 'N/A'
        try:
            date_formats = [
                '%Y-%m-%d %H:%M:%S.%f',
                '%Y-%m-%d %H:%M:%S',
                '%Y-%m-%d'
            ]
            
            for fmt in date_formats:
                try:
                    date_obj = datetime.strptime(str(date_str), fmt)
                    return date_obj.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            
            return 'N/A'
        except Exception:
            return 'N/A'

    def load_combined_data(self):
        for i in self.user_tree.get_children():
            self.user_tree.delete(i)

        users = list(self.db.users_collection.find())
        education_data = list(self.db.education_collection.find())
        internship_data = list(self.db.internships_collection.find())

        education_map = {str(edu.get('user_id')): edu for edu in education_data}
        internship_map = {str(intern.get('user_id')): intern for intern in internship_data}

        for index, user in enumerate(users, 1):
            user_id = str(user.get('_id'))
            edu_data = education_map.get(user_id, {})
            intern_data = internship_map.get(user_id, {})

            resume_status = "View PDF" if edu_data.get('resume') else "No Resume"

            item_id = self.user_tree.insert('', 'end', values=(
                index,  
                user.get('first_name', 'N/A'),
                user.get('middle_name', 'N/A'),
                user.get('last_name', 'N/A'),
                user.get('age', 'N/A'),
                user.get('gender', 'N/A'),
                user.get('contact_no', 'N/A'),
                user.get('email', 'N/A'),
                user.get('address', 'N/A'),
                edu_data.get('ssc_percentage', 'N/A'),
                edu_data.get('education_type', 'N/A'),
                edu_data.get('hsc_diploma_percentage', 'N/A'),
                edu_data.get('diploma_stream', 'N/A'),
                edu_data.get('degree_status', 'N/A'),
                edu_data.get('degree_type', 'N/A'),
                edu_data.get('degree_stream', 'N/A'),
                edu_data.get('degree_percentage', 'N/A'),
                edu_data.get('current_year', 'N/A'),
                edu_data.get('completion_year', 'N/A'),
                edu_data.get('skills', 'N/A'),
                resume_status,  
                intern_data.get('role', 'N/A'),
                intern_data.get('duration', 'N/A'),
                self.format_date(intern_data.get('joining_date', 'N/A')),
                self.format_date(intern_data.get('ending_date', 'N/A')),
                intern_data.get('internship_type', 'N/A'),
                intern_data.get('stipend_amount', 'N/A'),
                intern_data.get('stipend_frequency', 'N/A')
            ))
        
                
    def open_resume_from_education(self, edu_record):
        """Open the resume PDF from the education record."""
        if not edu_record or not edu_record.get('resume'):
            messagebox.showinfo("Resume", "No resume available for this education record.")
            return

        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
            temp_file = os.path.join(downloads_folder, 'temp_resume.pdf')
            
            resume_data = base64.b64decode(edu_record['resume'])
            
            with open(temp_file, 'wb') as f:
                f.write(resume_data)
            
            if os.name == 'nt':  
                os.startfile(temp_file)
            elif os.name == 'posix':  
                if sys.platform == 'darwin':  
                    subprocess.run(['open', temp_file])
                else:  # Linux
                    subprocess.run(['xdg-open', temp_file])
            else:
                webbrowser.open(temp_file)
                
        except Exception as e:
            messagebox.showerror("Error", f"Could not open resume: {str(e)}")

    def export_all_to_excel(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if not filename:
            return

        wb = openpyxl.Workbook()

        ws_combined = wb.create_sheet(title="User Details")
        combined_headers = [
            "Serial", "First Name", "Middle Name", "Last Name", 
            "Age", "Gender", "Contact", "Email", "Address",
            "SSC %", "Education Type", "HSC/Diploma %", 
            "Diploma Stream", "Degree Status", "Degree Type", "Degree Stream", 
            "Degree %", "Current Year", "Completion Year", "Skills",
            "Role", "Duration", "Joining Date", 
            "Ending Date", "Internship Type", "Stipend Amount", 
            "Stipend Frequency"
        ]
        ws_combined.append(combined_headers)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
        zip_filename = os.path.join(downloads_folder, f"resumes_{timestamp}.zip")

        with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
            users = list(self.db.users_collection.find())
            education_data = list(self.db.education_collection.find())
            internship_data = list(self.db.internships_collection.find())

            education_map = {str(edu.get('user_id')): edu for edu in education_data}
            internship_map = {str(intern.get('user_id')): intern for intern in internship_data}

            for index, user in enumerate(users, 1):
                user_id = str(user.get('_id'))
                edu_data = education_map.get(user_id, {})
                intern_data = internship_map.get(user_id, {})

                row_data = [
                    index, 
                    user.get('first_name', 'N/A'),
                    user.get('middle_name', 'N/A'),
                    user.get('last_name', 'N/A'),
                    user.get('age', 'N/A'),
                    user.get('gender', 'N/A'),
                    user.get('contact_no', 'N/A'),
                    user.get('email', 'N/A'),
                    user.get('address', 'N/A'),
                    edu_data.get('ssc_percentage', 'N/A'),
                    edu_data.get('education_type', 'N/A'),
                    edu_data.get('hsc_diploma_percentage', 'N/A'),
                    edu_data.get('diploma_stream', 'N/A'),
                    edu_data.get('degree_status', 'N/A'),
                    edu_data.get('degree_type', 'N/A'),
                    edu_data.get('degree_stream', 'N/A'),
                    edu_data.get('degree_percentage', 'N/A'),
                    edu_data.get('current_year', 'N/A'),
                    edu_data.get('completion_year', 'N/A'),
                    edu_data.get('skills', 'N/A'),
                    intern_data.get('role', 'N/A'),
                    intern_data.get('duration', 'N/A'),
                    self.format_date(intern_data.get('joining_date', 'N/A')),
                    self.format_date(intern_data.get('ending_date', 'N/A')),
                    intern_data.get('internship_type', 'N/A'),
                    intern_data.get('stipend_amount', 'N/A'),
                    intern_data.get('stipend_frequency', 'N/A')
                ]

                ws_combined.append(row_data)

                for col in [21, 22]:  
                    cell = ws_combined.cell(row=ws_combined.max_row, column=col+1)
                    if cell.value not in ['N/A', None]:
                        cell.number_format = 'YYYY-MM-DD'

                if edu_data.get('resume'):
                    try:
                        resume_data = base64.b64decode(edu_data['resume'])
                        user_name = f"{user.get('first_name', '')}_{user.get('last_name', '')}"
                        if not user_name.strip('_'):
                            user_name = f"user_{index}" 
                        pdf_filename = f"{user_name}.pdf"
                        
                        zipf.writestr(pdf_filename, resume_data)
                    except Exception as e:
                        print(f"Error saving resume for user {user_name}: {str(e)}")

        wb.remove(wb['Sheet'])

        wb.save(filename)
        
        messagebox.showinfo(
            "Export Successful", 
            f"Excel file exported to:\n{filename}\n\nResumes exported to:\n{zip_filename}"
        )